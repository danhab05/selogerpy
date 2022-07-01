# from src.getcontact import ScrapContacts
from flask import Flask, request
from flask_cors import CORS
from twilio.rest import Client
import urllib.request
import openpyxl
import requests as r
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
from time import sleep
from typing_extensions import Self
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from sympy import EX, false, true
from webdriver_manager.chrome import ChromeDriverManager
import pgeocode
from oauth2client.service_account import ServiceAccountCredentials
import gspread
from datetime import datetime
from datetime import date
from datetime import timedelta
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

app = Flask(__name__)
CORS(app)


class ScrapContacts:
    def __init__(self):
        options = Options()
        # options.headless =
        # options.add_argument("--window-size=1920,1080")
        # options.add_argument("--start-maximized")
        # options.add_argument("--headless")
        options.add_experimental_option(
            "excludeSwitches", ['enable-automation'])

        options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36")
        # options.add_argument("--remote-debugging-port=9222")

        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        options.add_argument("--log-level=OFF")
        self.driver = webdriver.Chrome(ChromeDriverManager(
        ).install(), service_log_path='NUL', options=options)

    def login(self):
        self.driver.get('https://myselogerpro.com/')
        while True:
            try:
                self.driver.find_element_by_xpath(
                    '/html/body/div[1]/div/div/div/div/div/div[2]/button[2]').click()
                break
            except Exception:
                print("waiting for login")
                pass
        sleep(0.2)
        while True:
            try:
                self.driver.find_element_by_xpath(
                    "/html/body/msp-root/msp-home-layout/div/div/div/div/div/section/login-login/div/div/form/div/div[2]/ui-input/div/label/input").send_keys("1091")
                sleep(0.2)
                passwd = "Lavieestbelle5782,"
                self.driver.find_element_by_xpath(
                    "/html/body/msp-root/msp-home-layout/div/div/div/div/div/section/login-login/div/div/form/div/div[3]/ui-input/div/label/input").send_keys(passwd)
                self.driver.find_element_by_xpath(
                    "/html/body/msp-root/msp-home-layout/div/div/div/div/div/section/login-login/div/div/form/div/div[4]/div/ui-button/div/button").click()
                break
            except Exception:
                pass
        print("Connected")
        sleep(7)

    def geContact(self, retake=false):
        continueC = True
        msg = ""

        nomi = pgeocode.Nominatim('fr')
        try:
            client = gspread.authorize(ServiceAccountCredentials.from_json_keyfile_name('key.json', [
                'https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']))
            sheet = client.open('seloger')
            sheet_instance = sheet.get_worksheet(0)
            if not retake:
                sheet_instance.clear()
        except Exception as e:
            print(e)
            pass
        self.driver.get("https://myselogerpro.com/contacts/acquereurs")
        print("Page acquereur")
        while True:
            try:
                self.driver.execute_script("window.scrollTo(0, 1000)")
                self.driver.find_element_by_xpath(
                    "/html/body/msp-root/msp-app-layout/main/cnt-contacts-outlet/div/cnt-contacts/div/div[2]/ui-page-size-selector/div/span/span[2]").click()
                self.driver.find_element_by_xpath(
                    "/html/body/msp-root/msp-app-layout/main/cnt-contacts-outlet/div/cnt-contacts/div/div[2]/ui-page-size-selector/div/div/div/div[3]").click()
                sleep(15)
                self.driver.execute_script("window.scrollTo(0, 0)")
                break
            except Exception:
                pass
        i = 1
        while True:
            try:
                print("Geting data...")
                dataLst = []
                self.driver.execute_script(
                    "window.scrollTo(0, document.body.scrollHeight);")
                # get data
                soup = BeautifulSoup(self.driver.page_source, "lxml")
                table = soup.find(
                    'cdk-table', {'class': 'cdk-table cnt-contacts-list__table cccl-table'})
                # , {'class': 'cdk-row cccl-table__row cccl-table-row cccl-table-row--highlighted cccl-table-row--clickable'})
                lines = table.find_all('cdk-row')
                for el in lines:
                    datestr = el.find(
                        "cdk-cell", {'class': 'cdk-cell cccl-table-row__cell cccl-table-row-cell date cdk-column-date ng-star-inserted'}).get_text()

                    dt1 = datestr
                    dt_obj = datetime.strptime(dt1, '%d/%m/%Y')
                    namestr = el.find(
                        "cdk-cell", {'class': 'cdk-cell cccl-table-row__cell cccl-table-row-cell cdk-column-name ng-star-inserted'}).get_text()

                    phonestr = el.find(
                        "cdk-cell", {'class': 'cdk-cell cccl-table-row__cell cccl-table-row-cell cdk-column-phone ng-star-inserted'}).get_text()

                    bienstr = el.find(
                        "cdk-cell", {'class': 'cdk-cell cccl-table-row__cell cccl-table-row-cell cdk-column-goods ng-star-inserted'}).get_text()

                    pricestr = el.find(
                        "cdk-cell", {'class': 'cdk-cell cccl-table-row__cell cccl-table-row-cell cdk-column-price ng-star-inserted'}).get_text()

                    surfacestr = el.find(
                        "cdk-cell", {'class': 'cdk-cell cccl-table-row__cell cccl-table-row-cell cdk-column-area ng-star-inserted'}).get_text()

                    zipCodestr = el.find(
                        "cdk-cell", {'class': 'cdk-cell cccl-table-row__cell cccl-table-row-cell cdk-column-location ng-star-inserted'}).get_text()

                    try:
                        city = nomi.query_postal_code(
                            int(zipCodestr)).place_name
                        if "," in city:
                            city = city.split(",")[0]
                    except Exception:
                        city = ""

                    dataLst.append([datestr, namestr, phonestr, bienstr,
                                   pricestr, surfacestr, zipCodestr, city])
                    todayweek = date.today()
                    todayweek -= timedelta(days=56)

                    newdate1 = datetime.strptime(str(todayweek), "%Y-%m-%d")
                    newdate2 = datetime.strptime(
                        str(dt_obj.date()), "%Y-%m-%d")
                    if newdate1 > newdate2:
                        continueC = False
                        break

                while True:
                    try:
                        entry = sheet_instance.ginsert_rows(
                            dataLst, len(sheet_instance.get_all_values()) + 1)
                        print("Added")
                        break
                    except Exception as e:
                        print(e)
                        pass

                if not continueC:
                    try:
                        self.driver.close()
                    except Exception as e:
                        print(e, "passed")
                        pass
                    break

                print(todayweek, str(dt_obj.date()))
                i += 1
                print("Next Page... - " + str(i))

                nav = soup.find("ul", {"class": "ui-pagination"})
                if len(nav.find_all("li")) == 11:
                    self.driver.find_element_by_xpath(
                        "/html/body/msp-root/msp-app-layout/main/cnt-contacts-outlet/div/cnt-contacts/div/div[2]/ui-pagination-old/ul/li[10]").click()
                else:
                    self.driver.find_element_by_xpath(
                        "/html/body/msp-root/msp-app-layout/main/cnt-contacts-outlet/div/cnt-contacts/div/div[2]/ui-pagination-old/ul/li[9]").click()
                # sleep(30)

                while BeautifulSoup(self.driver.page_source, 'lxml').find("div", {'class': "ui-loader__overlay"}) != None:
                    pass

            except Exception as e:
                print("sdqdqsd", e, "qsdqsdpqds")
                print("C'est la")
                if "id" in str(e) or "chrome not reachable" in str(e):
                    msg = "c"
                    break

        return "c"

    def close(self):
        self.driver.close()

    def sendEmail(self):
        try:
            client = gspread.authorize(ServiceAccountCredentials.from_json_keyfile_name('key.json', [
                'https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']))
            sheet = client.open('seloger')
            sheet_instance = sheet.get_worksheet(0)
        except Exception as e:
            pass
        link = "https://docs.google.com/spreadsheets/d/%s" % sheet.id
        try:
            fromaddr = "snowy27011@gmail.com"
            # toaddr = "mickaelsonego@gmail.com"
            toaddr = "danhabib011@gmail.com"

            msg = MIMEMultipart()
            msg['Subject'] = "Lien excel se loger"
            msg['From'] = fromaddr
            msg['To'] = toaddr
            html = f"""\
            <html>
            <body>
                <p>
                Bonjour voici le lien du excel se loger: <br>
                <a href="{link}">{link}</a> 
                </p>
            </body>
            </html>
            """
            part1 = MIMEText(html, "html")
            msg.attach(part1)
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(fromaddr, 'pbcygpldgslogvzf')
            server.sendmail(fromaddr, toaddr, msg.as_string())

        except Exception as e:
            print(e)
            pass


@app.route('/')
def index():
    return 'Ok'


@app.route('/getcontact')
def getContact():
    browser = ScrapContacts()
    browser.login()
    try:
        browser.geContact(retake=False)
    except Exception:
        browser.close()
        pass
    browser.sendEmail()
    return "ok"


@app.route("/sms")
def sendMsg():
    print('Geting excel')
    link = request.args.get("link")
    msg = request.args.get("msg")
    urllib.request.urlretrieve(link, "excel.xlsx")
    print("Done")
    dateSeLoger = 0
    name = 1
    phoneNumber = 2
    kind = 3
    price = 4
    surface = 5
    zipCode = 6
    city = 7
    wookbook = openpyxl.load_workbook("excel.xlsx")
    worksheet = wookbook.active
    for col in worksheet.iter_rows(1, worksheet.max_row):
        # print(col[1].value, col[2].value, col[3].value, col[4].value,)
        if not all([cell.value == None for cell in col]):
            try:
                newMsg = msg.replace("[Date seloger]", str(col[dateSeLoger].value)).replace("[Nom Prenom]", str(col[name].value)).replace("[Telephone]", str(col[phoneNumber].value)).replace(
                    "[type]", col[kind].value).replace("[prix]", col[price].value).replace("[surface]", col[surface].value).replace("[Code postal]", col[zipCode].value).replace("[ville]", col[city].value)
                number = str(col[phoneNumber].value)
                number = number.replace(" ", "")
                if "+" not in number:
                    number = "+33" + number

                try:
                    phone = col[phoneNumber].value
                    sms = newMsg
                    # link = f"https://trigger.macrodroid.com/95686f40-af84-472d-ab3b-95cf9a529709/sms?eau={phone}&&pain={sms}"
                    link = f"https://trigger.macrodroid.com/d5a0e1dd-7222-4f7c-8b32-aedd70bab2d6/sms?destinataire={phone}&&sms={sms}"
                    r.get(link)
                    # account_sid = 'ACc8c2f62d7f2b00775a667f13933f2e6f'
                    # auth_token = '7c36d2b8b60ffd3e8d783f87853241a0'
                    # client = Client(account_sid, auth_token)
                    # message = client.messages.create(
                    #     messaging_service_sid='MGca1c4db4a73fb12999bc81123d7d946d',
                    #     body=newMsg,
                    #     to='+330651598405'
                    #     # to=col[phoneNumber].value
                    # )
                    print("Sended")
                except Exception as e:
                    print(e)
                    pass
            except Exception as e:
                print(e)
                pass
    return "ok"


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=8099)
